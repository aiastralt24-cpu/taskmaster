#!/usr/bin/env python3
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from server import DB_PATH, connect, init_db, insert_many, normalize_task, now_iso  # noqa: E402

WORKBOOK = Path("/Users/bunny/Downloads/Astral_Video_Content_Tracker - F (1).xlsx")
SHEET = "Master Tracker"

STATUS_MAP = {
    "": "Concept Stage",
    "Not Started": "Concept Stage",
    "Concept Stage": "Concept Stage",
    "WIP": "In Production",
    "WRM to revert": "In Review",
    "Ready to go live": "Ready",
    "Completed": "Completed",
    "Went Live": "Published",
    "Published": "Published",
    "On Hold": "On Hold",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def format_length(value):
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        if int(value) == value:
            return f"{int(value)} Sec"
        return f"{value} Sec"
    text = clean(value)
    if text.isdigit():
        return f"{text} Sec"
    return text


def format_deadline(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def status_for(value):
    raw = clean(value)
    return STATUS_MAP.get(raw, raw if raw in set(STATUS_MAP.values()) else "Concept Stage")


def load_rows(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[SHEET]
    headers = [clean(cell.value) for cell in ws[4]]
    index = {name: i for i, name in enumerate(headers)}
    required = ["No.", "Brand", "Video Zone", "Video Title", "Length", "Hook – Main Subject", "Agency", "Status", "Primary Owner", "Secondary Owner", "Expected Completion Date"]
    missing = [name for name in required if name not in index]
    if missing:
        raise SystemExit(f"Missing workbook columns: {', '.join(missing)}")

    rows = []
    for excel_row in ws.iter_rows(min_row=5, values_only=True):
        title = clean(excel_row[index["Video Title"]])
        no_value = excel_row[index["No."]]
        if not title and not no_value:
            continue
        no = int(no_value) if isinstance(no_value, (int, float)) and no_value else len(rows) + 1
        zone = clean(excel_row[index["Video Zone"]])
        primary = clean(excel_row[index["Primary Owner"]])
        status = status_for(excel_row[index["Status"]])
        rows.append({
            "id": f"xlsx_{no:03d}",
            "no": no,
            "brand": clean(excel_row[index["Brand"]]),
            "bucket": zone,
            "zone": zone,
            "title": title,
            "length": format_length(excel_row[index["Length"]]),
            "hook": clean(excel_row[index["Hook – Main Subject"]]),
            "agency": clean(excel_row[index["Agency"]]),
            "status": status,
            "owner": primary,
            "primaryOwner": primary,
            "secondaryOwner": clean(excel_row[index["Secondary Owner"]]),
            "deadline": format_deadline(excel_row[index["Expected Completion Date"]]),
            "priority": "Medium",
            "notes": "",
        })
    return rows


def main():
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else WORKBOOK
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    init_db()
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = DB_PATH.with_suffix(DB_PATH.suffix + f".before-status-remap-{stamp}.bak")
    shutil.copy2(DB_PATH, backup)

    normalized = [normalize_task(row) for row in load_rows(workbook_path)]
    imported_at = now_iso()
    with connect() as conn:
        conn.execute("DELETE FROM tasks")
        conn.execute(
            "INSERT INTO activity_log(taskId,action,actorRole,changedAt,beforeJson,afterJson) VALUES(?,?,?,?,?,?)",
            ["import", "workbook status remap import", "admin", imported_at, "", f"{workbook_path}"],
        )
        insert_many(conn, normalized)
    print(f"Imported {len(normalized)} tasks")
    print(f"Backup: {backup}")


if __name__ == "__main__":
    main()
